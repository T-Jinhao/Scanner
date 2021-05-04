#!/usr/bin/python
# -*- coding:utf8 -*-
#author:Jinhao

import os
import datetime
from openpyxl import Workbook,load_workbook
from lib.color_output import *

class Report:
    def __init__(self, report, taskname, sheetname, banner, suc_msg='保存成功', err_msg='保存出错', lable=':'):
        '''
        保存xlsx格式的结果
        :param report: 报告内容,list
        :param taskname: 任务名兼文本名
        :param sheetname: xlsx工作台名称，用作分页
        :param banner: 工作页第一行信息表,list
        :param suc_msg: 保存成功输出文本
        :param err_msg: 保存失败输出文本
        :param lable: 用作文本切割
        '''
        self.report = report
        self.taskname = str(taskname)
        self.sheetname = sheetname
        self.banner = banner
        self.suc_msg = suc_msg
        self.err_msg = err_msg
        self.lable = lable
        self.Output = ColorOutput()

    def save(self):
        if self.report == []:
            print(self.Output.green('[ 保存输出结果-xlsx ] ') + self.Output.yellow('保存出错，报告为空'))
            return
        file = self.getAbsolutePath(self.taskname)
        isS = self.autoSetWorkbook(file)
        if isS:
            print(self.Output.green('[ 保存输出结果-xlsx ] ') + self.Output.fuchsia('分支：' + self.sheetname + self.suc_msg))
        else:
            print(self.Output.green('[ 保存输出结果-xlsx ] ') + self.Output.yellow('分支：' + self.sheetname + self.err_msg))
        return


    def getAbsolutePath(self, taskname, ext=".xlsx"):
        '''
        返回要读写文件的绝对路径
        :param taskname:
        :return:
        '''
        path = os.path.abspath(os.path.dirname(__file__))
        if platform.system() == 'Windows':
            interval = '\\'
        else:
            interval = '/'
        file = interval.join([path, taskname+ext])
        return file

    def autoSetWorkbook(self, file):
        '''
        自动获取工作台
        :param file:
        :return:
        '''
        if os.path.exists(file):   # 文件存在，使用追加
            isS = self.appendWorkbook(file)
        else:                      # 文件不存在，新建对象
            isS = self.newWorkbook(file)
        return isS

    def newWorkbook(self, file):
        '''
        新建对象
        :param file:
        :return:
        '''
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheetname
        ws.append(self.banner)
        try:
            for m in self.report:
                ws.append(self.dict2list(m))
            wb.save(file)
            isS = True
        except Exception as e:
            try:
                bakFile = self.getAbsolutePath(self.taskname + '_bak')
                wb.save(filename=bakFile)
                isS = True
            except:
                isS = False
        return isS

    def appendWorkbook(self, file):
        '''
        文本附加至已有文件
        :param file:
        :return:
        '''
        wb = load_workbook(file)
        sheets = wb.sheetnames
        if self.sheetname in sheets:
            ws = wb[self.sheetname]
            ws.append([])
            # today = datetime.datetime.today()
            # formatted_today = today.strftime('%y%m%d')
            # ws.append(["修改日期："+formatted_today])
        else:
            ws = wb.create_sheet(title=self.sheetname)
            ws.append(self.banner)
        try:
            for m in self.report:
                ws.append(self.dict2list(m))
            wb.save(file)
            isS = True
        except:
            try:
                bakFile = self.getAbsolutePath(self.taskname+'_bak')
                wb.save(filename=bakFile)
                isS = True
            except:
                isS = False
        return isS

    def dict2list(self, data):
        '''
        字典转数组
        :param data:
        :return:
        '''
        if type(data) == dict:
            DATA = [str(data[x]) for x in self.lable]
        else:
            DATA = [x for x in data.split(':')]
        return DATA






