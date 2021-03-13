#!/usr/bin/python
# -*- coding:utf8 -*-
#author:Jinhao

import os
import datetime
from openpyxl import Workbook,load_workbook
from lib.color_output import *

class Report:
    def __init__(self, report, taskname, sheetname, banner, suc_msg='保存成功', err_msg='保存出错', cut=':'):
        '''
        保存xlsx格式的结果
        :param report: 报告内容,list
        :param taskname: 任务名兼文本名
        :param sheetname: xlsx工作台名称，用作分页
        :param banner: 工作页第一行信息表,list
        :param suc_msg: 保存成功输出文本
        :param err_msg: 保存失败输出文本
        :param cut: 用作文本切割
        '''
        self.report = report
        self.taskname = taskname
        self.sheetname = sheetname
        self.banner = banner
        self.suc_msg = suc_msg
        self.err_msg = err_msg
        self.cut = cut
        self.Output = ColorOutput()

    def save(self):
        if self.report == []:
            print(self.Output.green('[ 保存输出结果-xlsx ] ') + self.Output.yellow('保存出错，报告为空'))
            return
        file = self.getAbsolutePath(self.taskname)
        isS = self.autoSetWorkbook(file)
        if isS:
            print(self.Output.green('[ 保存输出结果-xlsx ] ') + self.Output.fuchsia('分支：' + self.sheetname + self.suc_msg))
        else:
            print(self.Output.green('[ 保存输出结果-xlsx ] ') + self.Output.yellow('分支：' + self.sheetname + self.err_msg))
        return


    def getAbsolutePath(self, taskname, ext=".xlsx"):
        '''
        返回要读写文件的绝对路径
        :param taskname:
        :return:
        '''
        path = os.path.abspath(os.path.dirname(__file__))
        if platform.system() == 'Windows':
            interval = '\\'
        else:
            interval = '/'
        file = interval.join([path, taskname+ext])
        return file

    def autoSetWorkbook(self, file):
        '''
        自动获取工作台
        :param file:
        :return:
        '''
        if os.path.exists(file):   # 文件存在，使用追加
            isS = self.appendWorkbook(file)
        else:                      # 文件不存在，新建对象
            isS = self.newWorkbook(file)
        return isS

    def newWorkbook(self, file):
        '''
        新建对象
        :param file:
        :return:
        '''
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheetname
        ws.append(self.banner)
        try:
            for m in self.report:
                ws.append(self.str2list(m))
            wb.save(file)
            isS = True
        except:
            isS = False
        return isS

    def appendWorkbook(self, file):
        '''
        文本附加至已有文件
        :param file:
        :return:
        '''
        wb = load_workbook(file)
        sheets = wb.sheetnames
        if self.sheetname in sheets:
            ws = wb[self.sheetname]
            today = datetime.datetime.today()
            formatted_today = today.strftime('%y%m%d')
            ws.append(["修改日期："+formatted_today])
        else:
            ws = wb.create_sheet(title=self.sheetname)
            ws.append(self.banner)
        try:
            for m in self.report:
                ws.append(self.str2list(m))
            wb.save(file)
            isS = True
        except:
            isS = False
        return isS

    def str2list(self, text):
        '''
        字符串转数组
        :param text:
        :return:
        '''
        L = text.split(self.cut)
        return [x.strip() for x in L]




